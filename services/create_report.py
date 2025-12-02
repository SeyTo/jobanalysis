import os
from dataclasses import asdict
from datetime import datetime

import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side

from models.job import Job

COLUMNS = {
    "rating": "Rating",
    "company": "Company",
    "title": "Job Title",
    "desc": "Job Description",
    "location": "Location",
    "level_desc": "Job 'Level'",
    "url": "Job Apply Link",
    "company_desc": "Company Description",
    "company_url": "Company Site",
    "employees_num": "Number of Employees",
    "is_remote": "Remote",
}


def create_report(jobs: list[Job]) -> str:
    df = pd.DataFrame([asdict(job) for job in jobs])[list(COLUMNS.keys())]

    df.rename(columns=COLUMNS, inplace=True)
    df.sort_values(by=["Rating", "Company"], ascending=[False, True], inplace=True)

    time = datetime.now()
    name = f"job_report_{time.day}-{time.month}-{time.year}"

    script_dir = os.path.dirname(os.path.abspath(__file__))
    df.to_excel(f"{os.path.join(script_dir)}/../reports/{name}.xlsx", index=False)
    _post_creation_changes(name)

    return name


def _post_creation_changes(report_name: str):
    script_dir = os.path.dirname(os.path.abspath(__file__))

    wb = load_workbook(f"{os.path.join(script_dir)}/../reports/{report_name}.xlsx")
    ws = wb.active

    for idx, row in enumerate(ws.iter_rows()):
        paint_color = "FFE599"

        for cell in row:
            # fix rows
            cell.alignment = Alignment(wrap_text=True)
            if cell.column_letter == "D":
                ws.column_dimensions["D"].width = 100
            else:
                ws.column_dimensions[cell.column_letter].bestFit = True

            if idx == 0:
                continue
            # check if different color
            if cell.column_letter == "A":
                if cell.value >= 80:
                    paint_color = "B6D7A8"
                elif cell.value <= 40:
                    paint_color = "EA9999"

            cell.fill = PatternFill(fgColor=paint_color, fill_type="solid")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

    wb.save(f"{os.path.join(script_dir)}/../reports/{report_name}.xlsx")
